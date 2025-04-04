from src.docs.document_generator_interface import DocumentGeneratorInterface
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font

class ExcelDocumentGenerator(DocumentGeneratorInterface):
    def generate_document(self, document_data: dict, save_path: str):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Таблица"

        for column_name, column_dict in document_data["columns"].items():
            x_id = int(column_name.split("_")[-1])
            
            column = column_dict["column"]
            column_letter = get_column_letter(x_id+1)
            sheet.column_dimensions[column_letter].width = column_dict["width"]

            for value_dict in column:
                row_name, value = list(value_dict.items())[0]
                y_id = int(row_name)
                sheet.row_dimensions[y_id + 1].height = None
                
                
                excel_id = column_letter + str(y_id + 1)
                
                sheet[excel_id] = value
                sheet[excel_id].alignment = Alignment(wrapText=True)

                is_bordered = bool(value_dict["border"])
                if is_bordered:
                    border_style = Border(
                        left=Side(style="thin"),   # Левая граница
                        right=Side(style="thin"),  # Правая граница
                        top=Side(style="thin"),    # Верхняя граница
                        bottom=Side(style="thin")  # Нижняя граница
                    )

                    sheet[excel_id].border = border_style
                
                is_bold = bool(value_dict["bold"])
                if is_bold:
                    bold_font = Font(bold=True)
                    sheet[excel_id].font = bold_font
        
        for row_name, row_data in document_data["row_data"].items():
            y_id = int(row_name)
            sheet.row_dimensions[y_id + 1].height = row_data["height"]

        
        wb.save(save_path)
        return save_path